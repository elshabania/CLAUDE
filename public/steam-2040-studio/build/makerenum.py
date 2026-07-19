#!/usr/bin/env python3
"""One workbook: a renumbering sheet for every aggregation approach —
the 12 raced spatial methods at 3,150 & 3,000 zones, plus the app's built-in
methods (M1-M5, Distance change, Intrazonal demand) and every parameter
variation the optimizer races. Externals always numbered last."""
import json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = pathlib.Path(__file__).parent
D = json.loads((DIR / "renum-all.json").read_text())
A = json.loads((DIR / "renum-app.json").read_text())
PRE = json.loads((DIR / "preload-agg.json").read_text())["results"]
IDS = sorted(D.pop("__ids"))
EXT = [z for z in IDS if z >= 5995]

AR = Font(name="Arial", size=9.5)
BOLD = Font(name="Arial", size=9.5, bold=True)
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=12.5, bold=True)
SUB = Font(name="Arial", size=9, color="444444")
HFILL = PatternFill("solid", fgColor="1F3864")
MERGED = PatternFill("solid", fgColor="FFF2CC")
EXTF = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

FAMDESC = {
 "gehx": ("GEHX exact-guard", "Network-aware, error-guarded aggregation. Stage 1 merges zones sharing the same road attachment node (their demand loads at the identical point, so the assignment is provably unchanged - GEH 0 on every link). Stage 2, if the target needs more, moves the smallest-demand zones to the nearest surviving zone; a six-variant race confirmed this policy optimal. External stations are frozen (never merged)."),
 "nnd": ("NND demand-weighted", "Adjacent-first merging where inter-zone distances are inflated by the zones' demand, so high-demand zones resist merging and low-demand zones merge first."),
 "qtd": ("QTD demand quadtree", "Quadtree that keeps splitting the cell with the most DEMAND until roughly the target number of occupied leaves; every zone in a leaf merges. Balances demand per zone rather than area."),
 "nn": ("NN adjacent-first", "Greedy nearest-pair merging: repeatedly merge the closest zone clusters until the target count. Purely geometric."),
 "ward": ("WARD variance-minimising", "Hierarchical merging minimising d^2 * n1*n2/(n1+n2) - keeps clusters compact and size-balanced. Produces the same partition as NN at these fine levels (singleton merges dominate)."),
 "kmeans": ("KM k-means", "Lloyd's algorithm: exactly k compact clusters by iterative centroid assignment. Ignores adjacency and demand."),
 "kcenter": ("KC k-center", "Farthest-point coverage: picks centres minimising the worst zone-to-centre distance, then assigns zones to the nearest centre."),
 "hex": ("HEX hexagonal cells", "Fixed pointy-top hexagonal binning, cell size solved so occupied cells match the target; every zone in a cell merges."),
 "grid": ("GRID square cells", "Square-cell binning, size solved to the target count; every zone in a cell merges."),
 "quad": ("QT quadtree (area)", "Quadtree splitting the fullest AREA cell until ~target occupied leaves. Density-adaptive by count, not demand."),
 "bal": ("BAL size-balanced", "Ward-family criterion d^2 * (n1+n2) that resists mega-clusters. Same partition as NN at these fine levels."),
 "ring": ("RING rings x sectors", "Concentric rings x angular sectors around the network centre, occupancy-adaptive. Cannot hit exact zone counts; the achieved count is stated below."),
}
PROTOCOL = ("Scores: identical 800-origin pre-sampled trips assigned on full zones (demand-matched baseline) and on the aggregated zones, "
            "BPR rerouting, whole network, carrying links only. Renumbering: internal zones first, sequential from 0 with no gaps, sorted by the "
            "representative's original ID; the 6 external stations (orig 5995-6000, blue) always keep the LAST positions, in original order, never merged. "
            "Merged internal zones (amber, IS_REP=0) map to their representative's new ID.")

# ---- the app's built-in methods + the parameter variations the optimizer races ----
BASE_APP = {
 "m1": ("M1 land use + distance", "The app's Method 1: neighbouring zones merge only when their land-use classes are compatible and the merge stays within distance limits; candidate merges are ordered by a land-use/distance score and the sequence is cut at the app's zone-count target (default ~2,050). Region and sector boundary guards are always on."),
 "m2": ("M2 districts", "The app's Method 2: zones merge only within their administrative DISTRICT group; each district is condensed by the same fraction, derived from the zone-count target, so district structure is preserved."),
 "m3": ("M3 low OD", "The app's Method 3: the merge order internalises the least travel interaction first (weakest-interacting neighbours merge before strong ones). With no separately loaded OD matrix in the zoning engine the interaction proxy is used and the sequence is cut at the zone-count target (default ~2,050)."),
 "m4": ("M4 hybrid", "The app's Method 4: combines M1's land-use compatibility with M3's low-interaction ordering - merges must be land-use compatible AND weakly interacting. The eligible sequence exhausts before the target, so the achieved count is what the guards allow."),
 "m5": ("M5 size / density", "The app's Method 5: smallest / least-dense zones (by land-use SHAPEAREA) merge first, cutting at the zone-count target (default ~2,050)."),
 "distance": ("Distance change (geometry)", "The app's distance-change criterion: zones merge along a density-first, barrier-constrained sequence for as long as the predicted trip-distance change stays within the tolerance (default 1%) at the chosen confidence (default 99%). Road barriers at the chosen class (default collector and above) block merges across major roads. The zone count is solved by the criterion, not set as a target."),
 "demand": ("Intrazonal demand", "The app's intrazonal-demand criterion: replays the same merge sequence and cuts where the cumulative demand that would become intrazonal exceeds the budget (% of total trips)."),
}
VARNOTE = {
 "strict": "VARIATION - strict land-use matching: only identical land-use classes may merge (base allows compatible classes), so fewer merges pass the guard.",
 "dens":   "VARIATION - per-km2 scoring: land-use compatibility is scored per unit area (density-normalised) instead of absolute, changing the merge order.",
 "sd":     "VARIATION - strict matching AND per-km2 scoring combined.",
 "urb":    "VARIATION - urban/rural guard on: zones merge only within the same urban/rural class.",
 "metro":  "VARIATION - metro-accessibility guard on: zones merge only within the same metro-access class.",
 "frt":    "VARIATION - freight guard on: freight-generating zones are protected from merging across the freight boundary.",
 "lu":     "VARIATION - land-use class guard on top of districts: merges must also match land-use class.",
 "art":    "VARIATION - arterials barrier: only arterials and freeways block merges (base also treats collectors as barriers), relaxing contiguity so more merges are eligible.",
 "c95":    "VARIATION - 95% confidence: merges accepted at 95% (base 99%), a less conservative test that admits more merges.",
 "t05":    "VARIATION - 0.5% tolerance: tighter trip-distance-change budget, so the sequence cuts earlier (fewer merges).",
 "t2":     "VARIATION - 2% tolerance: looser budget, so the sequence runs further (more merges).",
}
# (sheet name, json key). Order = method family, base first, then its variations.
APP_ORDER = [
 ("M1", "m1"), ("M1 STRICT", "m1_strict"), ("M1 PER-KM2", "m1_dens"), ("M1 STRICT+KM2", "m1_sd"),
 ("M1 +URBAN", "m1_urb"), ("M1 +METRO", "m1_metro"),
 ("M2", "m2"), ("M2 +LANDUSE", "m2_lu"), ("M2 +URBAN", "m2_urb"), ("M2 +METRO", "m2_metro"),
 ("M2 +FREIGHT", "m2_frt"), ("M2 ART-BARRIER", "m2_art"),
 ("M3", "m3"), ("M3 +URBAN", "m3_urb"), ("M3 +METRO", "m3_metro"), ("M3 +FREIGHT", "m3_frt"), ("M3 ART-BARRIER", "m3_art"),
 ("M4", "m4"), ("M4 +URBAN", "m4_urb"), ("M4 +METRO", "m4_metro"), ("M4 +FREIGHT", "m4_frt"), ("M4 ART-BARRIER", "m4_art"),
 ("M5", "m5"), ("M5 +URBAN", "m5_urb"), ("M5 +METRO", "m5_metro"), ("M5 +FREIGHT", "m5_frt"), ("M5 ART-BARRIER", "m5_art"),
 ("DIST", "dist"), ("DIST CONF95", "dist_c95"), ("DIST TOL0.5", "dist_t05"), ("DIST TOL2", "dist_t2"), ("DIST ART-BARRIER", "dist_art"),
]

def pre_metrics_custom(key):
    m, t = key.rsplit("_", 1); t = int(t)
    for r in PRE:
        c = r.get("cfg", {})
        if m == "gehx" and c.get("gehx") and c.get("target") == t: return r
        if c.get("custom") == m and c.get("target") == t: return r
    return None

def pre_metrics_app(method):
    for r in PRE:
        c = r.get("cfg", {})
        if c.get("method") == method and not c.get("guards") and not c.get("sets") and not c.get("custom") and not c.get("gehx"):
            return r
    return None

def root_fn(pairs):
    rep = {a: b for a, b in pairs}
    def root(z):
        n = 0
        while z in rep and n < 60: z = rep[z]; n += 1
        return z
    return root

def met_text(met):
    return (f" · GEH<5: {met['geh5']:.2f}% · GEH<2: {met['pctG2']:.2f}% · links within 1%: {met['pctW1']:.1f}%"
            f" · VHT error: {met['vhtErr']:.2f}% · correlation: {met['corr']:.3f} · demand internalised: {met['internPct']:.1f}%"
            f" · runtime {met['rt']:.0f}s vs baseline {met['baseRt']:.0f}s")

wb = Workbook()
idx = wb.active; idx.title = "Index"
idx["A1"] = "Zone renumbering — every aggregation approach: 12 raced methods at 3,150 / 3,000 zones + all app methods and their variations"
idx["A1"].font = TITLE
idx["A2"] = (PROTOCOL + " App-method sheets (M1-M5, DIST, DEMAND): the zone count is the method's natural outcome (its own target or "
             "solver), not forced to 3,150/3,000; any app-proposed merge touching an external station is dropped in post-processing "
             "(count noted per sheet). Base app configurations carry assignment scores from the 79-run grid; parameter variations "
             "were not separately assignment-scored.")
idx["A2"].font = SUB; idx["A2"].alignment = Alignment(wrap_text=True)
idx.merge_cells("A2:F2"); idx.row_dimensions[2].height = 78
for c, h in enumerate(["Sheet", "Method", "Target", "Zones achieved", "GEH<5 (%)", "GEH<2 (%)"], 1):
    cell = idx.cell(4, c, h); cell.font = HDR; cell.fill = HFILL; cell.border = THIN

ORDER = [f"{m}_{t}" for t in (3150, 3000) for m in
         ("gehx","nnd","qtd","nn","ward","bal","kmeans","kcenter","hex","grid","quad","ring")]
SHEETN = {"gehx":"GEHX","nnd":"NND","qtd":"QTD","nn":"NN","ward":"WARD","bal":"BAL",
          "kmeans":"KM","kcenter":"KC","hex":"HEX","grid":"GRID","quad":"QT","ring":"RING"}
irow = 5

def write_sheet(name, fam, desc, pairs, met, target_lbl, extra_res="", title=None):
    """One renumbering sheet + its Index row. Returns achieved zone count."""
    global irow
    ext_set = set(EXT)
    kept = [p for p in pairs if p[0] not in ext_set and p[1] not in ext_set]
    ext_dropped = len(pairs) - len(kept)
    root0 = root_fn(kept)
    idset = set(IDS)
    # canonicalise: a representative must itself be a model zone — clusters whose
    # union-find root lies outside the model table re-root to their smallest
    # in-table member (the viewer's zone universe has 22 extra display zones)
    clus = {}
    for z in IDS: clus.setdefault(root0(z), []).append(z)
    canon = {}
    for r0, mem in clus.items():
        canon[r0] = r0 if r0 in idset else min(mem)
    def root(z): return canon[root0(z)]
    reps = sorted({root(z) for z in IDS})
    int_reps = [r for r in reps if r not in ext_set]
    new_of = {r: i for i, r in enumerate(int_reps)}
    for k, e in enumerate(sorted(EXT)): new_of[e] = len(int_reps) + k
    assert sorted(new_of.values()) == list(range(len(reps))), name
    members = {}
    for z in IDS: members[root(z)] = members.get(root(z), 0) + 1

    ws = wb.create_sheet(name)
    ws["A1"] = (title or f"{fam}{' @ target ' + target_lbl if target_lbl else ''}") + " — zone renumbering"; ws["A1"].font = TITLE
    ws["A2"] = "METHODOLOGY: " + desc; ws["A2"].font = SUB
    ws["A2"].alignment = Alignment(wrap_text=True); ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 46
    res = f"RESULTS: {len(reps):,} zones achieved (from 3,670)"
    if met: res += met_text(met)
    if ext_dropped: res += f" · {ext_dropped} proposed merge(s) touching external stations dropped"
    if extra_res: res += " · " + extra_res
    res += ". " + PROTOCOL
    ws["A3"] = res; ws["A3"].font = SUB
    ws["A3"].alignment = Alignment(wrap_text=True); ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 58

    for c, h in enumerate(["ORIG_ZONE_ID","NEW_ZONE_ID","REP_ORIG_ID","IS_REP","IS_EXTERNAL","MEMBERS_IN_NEW_ZONE"], 1):
        cell = ws.cell(5, c, h); cell.font = HDR; cell.fill = HFILL; cell.border = THIN
        cell.alignment = Alignment(horizontal="center")
    for i, oid in enumerate(IDS):
        r = root(oid); row = 6 + i
        is_ext = 1 if oid in ext_set else 0
        for c, v in enumerate([oid, new_of[r], r, 1 if r == oid else 0, is_ext, members[r]], 1):
            cell = ws.cell(row, c, v); cell.font = AR
        if is_ext:
            for c in range(1, 7): ws.cell(row, c).fill = EXTF
        elif r != oid:
            for c in range(1, 7): ws.cell(row, c).fill = MERGED
    for c, w in enumerate([15, 14, 14, 8, 12, 20], 1): ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A6"

    for c, v in enumerate([name, fam, target_lbl or "auto", len(reps),
                           round(met["geh5"], 2) if met else None,
                           round(met["pctG2"], 2) if met else None], 1):
        cell = idx.cell(irow, c, v); cell.font = AR; cell.border = THIN
    irow += 1
    return len(reps)

# ---- part 1: the 12 raced spatial methods at 3,150 and 3,000 ----
for key in ORDER:
    dat = D[key]; m, t = key.rsplit("_", 1)
    fam, desc = FAMDESC[m]
    write_sheet(f"{SHEETN[m]} {t}", fam, desc, dat["pairs"], pre_metrics_custom(key), t)

# ---- part 2: the app's built-in methods + variations ----
for name, key in APP_ORDER:
    dat = A.get(key)
    if not dat or "pairs" not in dat:
        continue
    base = key.split("_")[0] if key.split("_")[0] in BASE_APP else key
    if key.startswith("dist"): base = "distance"
    fam, desc = BASE_APP[base]
    var = key.split("_", 1)[1] if "_" in key else None
    if var and var in VARNOTE: desc = desc + " " + VARNOTE[var]
    met = pre_metrics_app(base) if not var else None
    extra = "" if not var else "this variation was not separately assignment-scored; the base configuration's scores are on its own sheet"
    # a variation identical to its base is still listed — say so honestly
    if var:
        bkey = key.split("_")[0]
        bdat = A.get(bkey if not key.startswith("dist") else "dist")
        if bdat and "pairs" in bdat and set(map(tuple, bdat["pairs"])) == set(map(tuple, dat["pairs"])):
            extra = ("this variation produced the IDENTICAL partition to the base configuration on this network"
                     + ("" if not extra else "; " + extra))
    write_sheet(name, fam, desc, dat["pairs"], met, None, extra,
                title=fam if not var else f"{fam} — variation: {name}")

# ---- intrazonal demand: documented as a no-op in this build ----
dem_ok = all((A.get(k) or {}).get("merged", 1) == 0 for k in ("dem_b05", "dem_b2", "dem_urb"))
fam, desc = BASE_APP["demand"]
desc += (" IN THIS BUILD the zoning engine has no separately loaded OD matrix (demand lives in the assignment engine), so the "
         "criterion cannot price any merge and performs NO merges at every budget tested (0.5%, 1%, 2%) and with the urban/rural "
         "guard - the 79-run comparison grid found the same. The table below is therefore the identity renumbering: every zone "
         "keeps its own representative and only the IDs compact to 0..3,669 with the 6 external stations last.")
write_sheet("DEMAND", fam, desc, [], None, None,
            "no merges at any tested budget (0.5 / 1 / 2%) or guard; identity renumbering shown")

for c, w in enumerate([16, 26, 9, 14, 11, 11], 1): idx.column_dimensions[get_column_letter(c)].width = w
idx.freeze_panes = "A5"

out = DIR / "Zone_Renumbering_All_Methods_3150_3000.xlsx"
wb.save(out)
print("sheets:", len(wb.sheetnames), "· size:", out.stat().st_size // 1024, "KB")
