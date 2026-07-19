#!/usr/bin/env python3
"""One workbook: every distinct aggregation method ONCE, all driven to ~3,000
zones. No repeated methods (NN/WARD/BAL collapse to one sheet - identical
partitions at this level) and no multiple aggregation levels. Externals last."""
import json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR = pathlib.Path(__file__).parent
D = json.loads((DIR / "renum-all.json").read_text())      # raced methods at exact 3000
T = json.loads((DIR / "renum-3000.json").read_text())     # app methods + tuned ring at ~3000, freshly scored
PRE = json.loads((DIR / "preload-agg.json").read_text())["results"]
IDS = sorted(D.pop("__ids"))
EXT = [z for z in IDS if z >= 5995]

AR = Font(name="Arial", size=9.5)
HDR = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE = Font(name="Arial", size=12.5, bold=True)
SUB = Font(name="Arial", size=9, color="444444")
HFILL = PatternFill("solid", fgColor="1F3864")
MERGED = PatternFill("solid", fgColor="FFF2CC")
EXTF = PatternFill("solid", fgColor="DDEBF7")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

PROTOCOL = ("Scores: identical 800-origin pre-sampled trips assigned on full zones (demand-matched baseline) and on the aggregated zones, "
            "BPR rerouting, whole network, carrying links only. Renumbering: internal zones first, sequential from 1 with no gaps (no zone is numbered 0), sorted by the "
            "representative's original ID; the 6 external stations (orig 5995-6000, blue) always keep the LAST positions, in original order, never merged. "
            "Merged internal zones (amber, IS_REP=0) map to their representative's new ID.")

def pre_metrics(kind, key):
    for r in PRE:
        c = r.get("cfg", {})
        if kind == "gehx" and c.get("gehx") and c.get("target") == 3000: return r
        if kind == "custom" and c.get("custom") == key and c.get("target") == 3000: return r
    return None

def met_text(met):
    return (f" · GEH<5: {met['geh5']:.2f}% · GEH<2: {met['pctG2']:.2f}% · links within 1%: {met['pctW1']:.1f}%"
            f" · VHT error: {met['vhtErr']:.2f}% · correlation: {met['corr']:.3f} · demand internalised: {met['internPct']:.1f}%"
            f" · runtime {met['rt']:.0f}s vs baseline {met['baseRt']:.0f}s")

def root_fn(pairs):
    rep = {a: b for a, b in pairs}
    def root(z):
        n = 0
        while z in rep and n < 60: z = rep[z]; n += 1
        return z
    return root

wb = Workbook()
idx = wb.active; idx.title = "Index"
idx["A1"] = "Zone renumbering — every distinct method once, all at ~3,000 zones"; idx["A1"].font = TITLE
idx["A2"] = (PROTOCOL + " Every sheet is one method at one level (~3,000 zones): the raced spatial methods hit 3,000 exactly; "
             "the app methods were driven to a 3,000-zone target (an option added to the app's target selector for this export); "
             "Distance uses the tolerance whose solved count lands nearest 3,000; RING's target was tuned by probing (it cannot hit "
             "exact counts). WARD and BAL are not repeated - they produce the IDENTICAL partition to NN at this level (noted on the NN "
             "sheet). All scores here use the same protocol, freshly run for the app methods at this level.")
idx["A2"].font = SUB; idx["A2"].alignment = Alignment(wrap_text=True)
idx.merge_cells("A2:F2"); idx.row_dimensions[2].height = 88
for c, h in enumerate(["Sheet", "Method", "Zones achieved", "GEH<5 (%)", "GEH<2 (%)", "Links within 1% (%)"], 1):
    cell = idx.cell(4, c, h); cell.font = HDR; cell.fill = HFILL; cell.border = THIN
irow = 5

def write_sheet(name, fam, desc, pairs, met, extra_res=""):
    global irow
    ext_set = set(EXT)
    kept = [p for p in pairs if p[0] not in ext_set and p[1] not in ext_set]
    ext_dropped = len(pairs) - len(kept)
    root0 = root_fn(kept)
    idset = set(IDS)
    # canonicalise: a representative must itself be a model zone — clusters whose
    # union-find root lies outside the model table re-root to their smallest
    # in-table member (the viewer's zone universe has 22 extra display zones)
    clus = {}
    for z in IDS: clus.setdefault(root0(z), []).append(z)
    canon = {r0: (r0 if r0 in idset else min(mem)) for r0, mem in clus.items()}
    def root(z): return canon[root0(z)]
    reps = sorted({root(z) for z in IDS})
    int_reps = [r for r in reps if r not in ext_set]
    new_of = {r: i + 1 for i, r in enumerate(int_reps)}   # 1-based: no zone gets ID 0
    for k, e in enumerate(sorted(EXT)): new_of[e] = len(int_reps) + k + 1
    assert sorted(new_of.values()) == list(range(1, len(reps) + 1)), name
    members = {}
    for z in IDS: members[root(z)] = members.get(root(z), 0) + 1

    ws = wb.create_sheet(name)
    ws["A1"] = f"{fam} @ ~3,000 zones — zone renumbering"; ws["A1"].font = TITLE
    ws["A2"] = "METHODOLOGY: " + desc; ws["A2"].font = SUB
    ws["A2"].alignment = Alignment(wrap_text=True); ws.merge_cells("A2:F2"); ws.row_dimensions[2].height = 52
    res = f"RESULTS: {len(reps):,} zones achieved (from 3,670)"
    if met: res += met_text(met)
    if ext_dropped: res += f" · {ext_dropped} proposed merge(s) touching external stations dropped"
    if extra_res: res += " · " + extra_res
    res += ". " + PROTOCOL
    ws["A3"] = res; ws["A3"].font = SUB
    ws["A3"].alignment = Alignment(wrap_text=True); ws.merge_cells("A3:F3"); ws.row_dimensions[3].height = 58

    for c, h in enumerate(["ORIG_ZONE_ID","NEW_ZONE_ID","REP_ORIG_ID","IS_REP","IS_EXTERNAL","MEMBERS_IN_NEW_ZONE"], 1):
        cell = ws.cell(5, c, h); cell.font = HDR; cell.fill = HFILL; cell.border = THIN
        cell.alignment = Alignment(horizontal="center")
    for i, oid in enumerate(IDS):
        r = root(oid); row = 6 + i
        is_ext = 1 if oid in ext_set else 0
        for c, v in enumerate([oid, new_of[r], r, 1 if r == oid else 0, is_ext, members[r]], 1):
            ws.cell(row, c, v).font = AR
        if is_ext:
            for c in range(1, 7): ws.cell(row, c).fill = EXTF
        elif r != oid:
            for c in range(1, 7): ws.cell(row, c).fill = MERGED
    for c, w in enumerate([15, 14, 14, 8, 12, 20], 1): ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A6"

    for c, v in enumerate([name, fam, len(reps),
                           round(met["geh5"], 2) if met else None,
                           round(met["pctG2"], 2) if met else None,
                           round(met["pctW1"], 1) if met else None], 1):
        cell = idx.cell(irow, c, v); cell.font = AR; cell.border = THIN
    irow += 1

# ---- raced spatial methods at exactly 3,000 ----
RACED = [
 ("GEHX", "gehx_3000", "gehx", None,
  "GEHX exact-guard: Stage 1 merges zones sharing the same road attachment node (provably identical assignment, GEH 0 on every link); "
  "Stage 2 moves the smallest-demand zones to the nearest surviving zone until the target. A six-variant race confirmed this policy optimal. "
  "External stations are frozen."),
 ("NND", "nnd_3000", "custom", "nnd",
  "NND demand-weighted (an enhancement of NN): adjacent-first merging with inter-zone distances inflated by demand, so high-demand zones "
  "resist merging and low-demand zones merge first."),
 ("QTD", "qtd_3000", "custom", "qtd",
  "QTD demand quadtree (an enhancement of QT): keeps splitting the cell with the most DEMAND until ~target occupied leaves; every zone in a "
  "leaf merges. Balances demand per zone rather than area."),
 ("NN", "nn_3000", "custom", "nn",
  "NN adjacent-first: greedy nearest-pair merging until the target count. Purely geometric. NOT REPEATED: WARD (variance-minimising) and "
  "BAL (size-balanced) produce the IDENTICAL partition at this level - singleton merges dominate, so all three hierarchical criteria pick "
  "the same pairs; one sheet covers all three."),
 ("KM", "kmeans_3000", "custom", "kmeans",
  "KM k-means: exactly k compact clusters by iterative centroid assignment. Ignores adjacency and demand."),
 ("KC", "kcenter_3000", "custom", "kcenter",
  "KC k-center: farthest-point coverage - centres minimise the worst zone-to-centre distance, zones assigned to the nearest centre."),
 ("HEX", "hex_3000", "custom", "hex",
  "HEX hexagonal cells: fixed pointy-top hexagonal binning, cell size solved so occupied cells match the target; every zone in a cell merges."),
 ("GRID", "grid_3000", "custom", "grid",
  "GRID square cells: square-cell binning, size solved to the target count; every zone in a cell merges."),
 ("QT", "quad_3000", "custom", "quad",
  "QT quadtree (area): splits the fullest AREA cell until ~target occupied leaves. Density-adaptive by count, not demand."),
]
for name, key, kind, ck, desc in RACED:
    write_sheet(name, {"GEHX":"GEHX exact-guard","NND":"NND demand-weighted","QTD":"QTD demand quadtree","NN":"NN adjacent-first (= WARD = BAL)",
                       "KM":"KM k-means","KC":"KC k-center","HEX":"HEX hexagonal cells","GRID":"GRID square cells","QT":"QT quadtree (area)"}[name],
                desc, D[key]["pairs"], pre_metrics(kind, ck))

# ---- tuned RING + app methods at the injected 3,000 target, freshly scored ----
APP = [
 ("RING", "ring_t", "RING rings x sectors",
  "RING rings x sectors: concentric rings x angular sectors around the network centre, occupancy-adaptive. It cannot hit exact zone counts, "
  "so the ring/sector target was tuned by probing (3,200-3,650) and the setting whose ACHIEVED count lands nearest 3,000 was kept.",
  ""),
 ("M1", "m1_3000", "M1 land use + distance",
  "The app's Method 1: neighbouring zones merge only when their land-use classes are compatible and within distance limits, ordered by a "
  "land-use/distance score. Driven to a 3,000-zone target (an option added to the app's target selector for this export). Region and sector "
  "guards on (app defaults).", ""),
 ("M2", "m2_3000", "M2 districts",
  "The app's Method 2: zones merge only within their administrative DISTRICT group, each district condensed by the same fraction derived "
  "from the 3,000-zone target - per-district rounding makes the achieved count approximate.", ""),
 ("M3", "m3_3000", "M3 low OD",
  "The app's Method 3: the merge order internalises the least travel interaction first; sequence cut at the 3,000-zone target.", ""),
 ("M4", "m4_3000", "M4 hybrid",
  "The app's Method 4: merges must be land-use compatible AND weakly interacting (M1 x M3). At ~3,000 zones the guarded sequence has ample "
  "headroom (it exhausts at ~2,129), so the target is reached exactly in the app's zone universe.", ""),
 ("M5", "m5_3000", "M5 size / density",
  "The app's Method 5: smallest / least-dense zones (by land-use SHAPEAREA) merge first; sequence cut at the 3,000-zone target.", ""),
 ("DIST", "dist_3000", "Distance change (geometry)",
  "The app's distance-change criterion: zones merge along a density-first, barrier-constrained sequence while the predicted trip-distance "
  "change stays within tolerance at 99% confidence. The zone count is solved, not targeted: 0.5% tolerance is the app setting whose solved "
  "count lands nearest 3,000.", ""),
]
for name, key, fam, desc, extra in APP:
    dat = T.get(key)
    if not dat or "pairs" not in dat:
        print("MISSING:", key); continue
    write_sheet(name, fam, desc, dat["pairs"], dat.get("score"), extra)

# ---- intrazonal demand: documented no-op ----
write_sheet("DEMAND", "Intrazonal demand",
            "The app's intrazonal-demand criterion would cut the merge sequence where cumulative intrazonalised trips exceed a budget (% of "
            "total). IN THIS BUILD the zoning engine has no separately loaded OD matrix (demand lives in the assignment engine), so the "
            "criterion cannot price any merge and performs NO merges at any budget - a ~3,000-zone system cannot be produced with it. The "
            "table is the identity renumbering: IDs compact to 1..3,670 with the 6 external stations last.",
            [], None, "no merges possible in this build; identity renumbering shown")

for c, w in enumerate([10, 30, 14, 11, 11, 16], 1): idx.column_dimensions[get_column_letter(c)].width = w
idx.freeze_panes = "A5"

out = DIR / "Zone_Renumbering_All_Methods_at_3000.xlsx"
wb.save(out)
print("sheets:", len(wb.sheetnames), "· size:", out.stat().st_size // 1024, "KB")
